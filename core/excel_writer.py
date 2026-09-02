from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.data_source import (
    NumData,
    NumVal,
    StrData,
    StrRef,
    StrVal,
)
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    CharacterProperties,
    Paragraph,
    ParagraphProperties,
    RichTextProperties,
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


# ============================================================
# GENEL STİLLER
# ============================================================

HEADER_FILL = PatternFill(
    "solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

TITLE_FILL = PatternFill(
    "solid",
    fgColor="D9EAF7",
)

THIN_BORDER = Border(
    left=Side(
        style="thin",
        color="D9E2F3",
    ),
    right=Side(
        style="thin",
        color="D9E2F3",
    ),
    top=Side(
        style="thin",
        color="D9E2F3",
    ),
    bottom=Side(
        style="thin",
        color="D9E2F3",
    ),
)

OK_FILL = PatternFill(
    "solid",
    fgColor="E2F0D9",
)

CONTROL_FILL = PatternFill(
    "solid",
    fgColor="FFF2CC",
)

ERROR_FILL = PatternFill(
    "solid",
    fgColor="FCE4D6",
)

BEST_FILL = PatternFill(
    "solid",
    fgColor="E2F0D9",
)

BEST_FONT = Font(
    color="006100",
    bold=True,
)

WEEKEND_FILL = PatternFill(
    "solid",
    fgColor="FFE699",
)

STATUS_LABELS = {
    "OK": "DOĞRU",
    "KONTROL": "KONTROL GEREKLİ",
    "ERROR": "HATA",
}


# ============================================================
# ÜRÜN RENKLERİ
# ============================================================

USD_FILL = PatternFill(
    "solid",
    fgColor="DDEBF7",
)

EUR_FILL = PatternFill(
    "solid",
    fgColor="FCE4D6",
)

XAU_FILL = PatternFill(
    "solid",
    fgColor="FFF2CC",
)

USD_HEADER_FILL = PatternFill(
    "solid",
    fgColor="4472C4",
)

EUR_HEADER_FILL = PatternFill(
    "solid",
    fgColor="ED7D31",
)

XAU_HEADER_FILL = PatternFill(
    "solid",
    fgColor="BF9000",
)

PRODUCT_LINE_COLORS = {
    "USD": "4472C4",
    "EUR": "ED7D31",
    "XAU": "FFC000",
}


# ============================================================
# SAĞLAYICI RENKLERİ
# ============================================================

PROVIDER_COLORS = {
    "Akbank": "ECCACA",
    "Albaraka Türk": "DFF1E4",
    "Alternatif Bank": "DECDEA",
    "Altınkaynak": "F3F0DD",
    "Anadolubank": "CDE5EA",
    "CEPTETEB": "F1DFE9",
    "Denizbank": "D3ECCA",
    "DestekBank": "E0DFF1",
    "Dünya Katılım": "EAD7CD",
    "Emlak Katılım": "DDF3EB",
    "Enpara": "E7CDEA",
    "Fibabanka": "EDF1DF",
    "Garanti BBVA": "CADBEC",
    "Getirfinans": "F1DFE3",
    "Halkbank": "CDEACF",
    "Harem": "E5DDF3",
    "Hayat Finans": "EAE0CD",

    # Kaynakta varsa mutlaka güncel sayfada gösterilir.
    "Hepsipay": "DFF1F1",

    "HSBC": "ECCAE4",
    "ING Bank": "E7F1DF",
    "İş Bankası": "CDD2EA",
    "Kapalıçarşı": "F3E0DD",
    "Kuveyt Türk": "CDEAD9",
    "Merkez Bankası": "ECDFF1",
    "Misyon Bank": "ECECCA",
    "Odacı": "DFECF1",
    "Odeabank": "EACDD9",

    # Kaynakta varsa mutlaka güncel sayfada gösterilir.
    "Papara": "E0F3DD",

    "QNB Finansbank": "D2CDEA",
    "TOM Bank Hadi": "F1E7DF",
    "Türkiye Finans": "CAECE4",
    "Vakıf Katılım": "F1DFF0",
    "Vakıfbank": "E0EACD",
    "Venüs": "DDE5F3",

    # Akbank'tan belirgin farklı mor.
    "Yapıkredi": "D9C2E9",

    "Ziraat Bankası": "DFF1E3",
    "Ziraat Dinamik": "DBCAEC",
    "Ziraat Katılım": "F1EDDF",
}

DEFAULT_PROVIDER_COLOR = "E8EDF3"


# ============================================================
# TEKİL BANKA GRAFİK ARKA PLANLARI
# ============================================================

BANK_CHART_COLORS = {
    "Garanti BBVA": "E2F0D9",
    "Akbank": "FDE2E2",
    "Yapıkredi": "E4D7F5",
    "Ziraat Bankası": "FFF2CC",
    "İş Bankası": "DDEBF7",
}

DEFAULT_BANK_CHART_COLOR = "E8EDF3"


# ============================================================
# 5 BANKA KARŞILAŞTIRMA ÇİZGİ RENKLERİ
# ============================================================

BANK_LINE_COLORS = {
    "Garanti BBVA": "70AD47",
    "Akbank": "C00000",
    "Yapıkredi": "7030A0",
    "Ziraat Bankası": "FFC000",
    "İş Bankası": "4472C4",
}


# ============================================================
# ÜRÜNLER
# ============================================================

PRODUCT_ORDER = {
    "USD": 0,
    "EUR": 1,
    "XAU": 2,
}

PRODUCT_NAMES = {
    "USD": "DOLAR",
    "EUR": "EURO",
    "XAU": "GRAM ALTIN",
}

TARGET_BANKS = [
    "Garanti BBVA",
    "Akbank",
    "Yapıkredi",
    "Ziraat Bankası",
    "İş Bankası",
]

MONTH_NAMES = {
    1: "Ocak",
    2: "Şubat",
    3: "Mart",
    4: "Nisan",
    5: "Mayıs",
    6: "Haziran",
    7: "Temmuz",
    8: "Ağustos",
    9: "Eylül",
    10: "Ekim",
    11: "Kasım",
    12: "Aralık",
}


# ============================================================
# GRAFİK YARDIMCI VERİLERİ
#
# AA sütunundan başlıyor.
#
# ÖNEMLİ:
# Bu sütunları GİZLEMİYORUZ.
#
# Çünkü Excel bazı sürümlerde gizli hücrelerden gelen grafik
# verilerini göstermeyebiliyor.
#
# AA sütunu zaten normal görünümün çok sağında kaldığı için
# kullanıcıyı rahatsız etmiyor.
# ============================================================

HELPER_START_COL = 27  # AA


# ============================================================
# YARDIMCI FONKSİYONLAR
# ============================================================

def _provider_fill(provider: str | None):
    provider = (
        provider or ""
    ).strip()

    color = PROVIDER_COLORS.get(
        provider,
        DEFAULT_PROVIDER_COLOR,
    )

    return PatternFill(
        "solid",
        fgColor=color,
    )


def _product_fill(code: str | None):
    code = (
        code or ""
    ).strip().upper()

    if code == "USD":
        return USD_FILL

    if code == "EUR":
        return EUR_FILL

    if code == "XAU":
        return XAU_FILL

    return PatternFill(
        fill_type=None
    )


def _parse_dt(
    value: str | None,
) -> datetime | None:

    if not value:
        return None

    value = str(value).strip()

    if not value:
        return None

    # ISO format
    try:
        dt = datetime.fromisoformat(
            value.replace(
                "Z",
                "+00:00",
            )
        )

        # Excel için timezone bilgisini kaldırıyoruz,
        # ancak duvar saatini koruyoruz.
        if dt.tzinfo is not None:
            dt = dt.replace(
                tzinfo=None
            )

        return dt

    except (
        ValueError,
        TypeError,
    ):
        pass

    formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(
                value,
                fmt,
            )
        except ValueError:
            continue

    return None


def _to_float(value):

    if value in (
        None,
        "",
    ):
        return None

    if isinstance(
        value,
        (int, float),
    ):
        return float(value)

    text = str(
        value
    ).strip()

    if not text:
        return None

    text = (
        text
        .replace("%", "")
        .replace("\u00a0", "")
        .strip()
    )

    try:
        return float(text)

    except ValueError:
        pass

    # Türkçe sayı biçimi:
    # 1.234,56 -> 1234.56
    if (
        "," in text
        and "." in text
    ):

        if (
            text.rfind(",")
            >
            text.rfind(".")
        ):
            text = (
                text
                .replace(".", "")
                .replace(",", ".")
            )
        else:
            text = text.replace(
                ",",
                "",
            )

    elif "," in text:
        text = text.replace(
            ",",
            ".",
        )

    try:
        return float(text)

    except ValueError:
        return None


def read_history(
    path: str | Path,
) -> list[dict]:

    path = Path(
        path
    )

    if not path.exists():
        return []

    with path.open(
        "r",
        encoding="utf-8-sig",
        newline="",
    ) as handle:

        return list(
            csv.DictReader(
                handle
            )
        )


def _style_header(
    ws,
    row: int,
    start_col: int,
    end_col: int,
):

    for col in range(
        start_col,
        end_col + 1,
    ):

        cell = ws.cell(
            row=row,
            column=col,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _apply_table(
    ws,
    start_row: int,
    end_row: int,
    end_col: int,
    name: str,
):

    if end_row <= start_row:
        return

    ref = (
        f"A{start_row}:"
        f"{get_column_letter(end_col)}"
        f"{end_row}"
    )

    table = Table(
        displayName=name,
        ref=ref,
    )

    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )

    ws.add_table(
        table
    )


def _set_widths(
    ws,
    widths,
):

    for column, width in widths.items():
        ws.column_dimensions[
            column
        ].width = width


def _status_fill(
    status,
):

    if status == "ERROR":
        return ERROR_FILL

    if status == "KONTROL":
        return CONTROL_FILL

    return OK_FILL


def _display_status(
    status,
):

    raw = (
        status or ""
    ).strip()

    return STATUS_LABELS.get(
        raw,
        raw,
    )


# ============================================================
# TARİH ETİKETİ
#
# Hafta sonlarında sadece etiket değişir.
#
# 28.08
# 29.08 Cmt
# 30.08 Paz
# 31.08
#
# Çizginin veya noktanın rengi değişmez.
# ============================================================

def _date_label(
    dt: datetime,
    include_time=False,
):

    if dt.weekday() == 5:
        suffix = " Cmt"

    elif dt.weekday() == 6:
        suffix = " Paz"

    else:
        suffix = ""

    label = (
        dt.strftime("%d.%m")
        + suffix
    )

    if include_time:
        label += (
            " "
            + dt.strftime("%H:%M")
        )

    return label


# ============================================================
# AYNI GÜNDEKİ İLK RUN
# ============================================================

def _daily_history(
    history: list[dict],
) -> list[dict]:

    first_run_by_day = {}

    for row in history:

        raw = row.get(
            "run_at"
        )

        dt = _parse_dt(
            raw
        )

        if (
            not raw
            or not dt
        ):
            continue

        day = dt.date()

        existing_raw = (
            first_run_by_day
            .get(day)
        )

        if existing_raw is None:

            first_run_by_day[
                day
            ] = raw

            continue

        existing_dt = _parse_dt(
            existing_raw
        )

        if (
            existing_dt is None
            or dt < existing_dt
        ):

            first_run_by_day[
                day
            ] = raw

    selected_runs = set(
        first_run_by_day.values()
    )

    result = []

    for row in history:

        if (
            row.get("run_at")
            in selected_runs
        ):
            result.append(
                row
            )

    return result


# ============================================================
# GERÇEK SON RUN
# ============================================================

def _latest_run_rows(
    history,
):

    valid = []

    for row in history:

        dt = _parse_dt(
            row.get("run_at")
        )

        if dt:
            valid.append(
                (
                    dt,
                    row,
                )
            )

    if not valid:
        return None, []

    latest_dt = max(
        dt
        for dt, _ in valid
    )

    # Aynı datetime'a ait tüm satırları al.
    latest_rows = []

    latest_raw = None

    for dt, row in valid:

        if dt == latest_dt:

            latest_rows.append(
                row
            )

            if latest_raw is None:
                latest_raw = (
                    row.get("run_at")
                )

    return (
        latest_raw,
        latest_rows,
    )


# ============================================================
# PROVIDER MAP
# ============================================================

def _provider_map(
    rows: Iterable[dict],
):

    result = {}

    for row in rows:

        provider = (
            row.get(
                "provider",
                "",
            )
            .strip()
        )

        code = (
            row.get(
                "code",
                "",
            )
            .strip()
            .upper()
        )

        if (
            not provider
            or not code
        ):
            continue

        result.setdefault(
            provider,
            {},
        )[code] = row

    return result


# ============================================================
# GEÇMİŞ SIRALAMA
# ============================================================

def _history_sort_key(
    row,
):

    dt = (
        _parse_dt(
            row.get("run_at")
        )
        or datetime.min
    )

    provider = (
        row.get("provider")
        or ""
    ).strip().casefold()

    code = (
        row.get("code")
        or ""
    ).strip().upper()

    return (
        dt,
        provider,
        PRODUCT_ORDER.get(
            code,
            99,
        ),
    )


# ============================================================
# MAKAS DEĞERİ
# ============================================================

def _get_spread_values(
    row,
):

    buy = _to_float(
        row.get("buy")
    )

    sell = _to_float(
        row.get("sell")
    )

    spread = _to_float(
        row.get("spread")
    )

    spread_pct = _to_float(
        row.get("spread_pct")
    )

    if (
        spread is None
        and buy is not None
        and sell is not None
    ):
        spread = (
            sell - buy
        )

    if (
        spread_pct is None
        and spread is not None
        and buy not in (
            None,
            0,
        )
    ):
        spread_pct = (
            spread / buy
        ) * 100

    return (
        buy,
        sell,
        spread,
        spread_pct,
    )


# ============================================================
# GRAFİK CACHE
#
# Excel Protected View / önizleme durumlarında grafiğin
# veri noktalarını korumaya yardımcı olur.
# ============================================================

def _cache_line_chart(
    chart,
    categories,
    series_values,
):

    try:

        for (
            series,
            (
                title,
                values,
            ),
        ) in zip(
            chart.series,
            series_values,
        ):

            numeric_points = []

            for index, value in enumerate(
                values
            ):

                if value is None:
                    continue

                numeric_points.append(
                    NumVal(
                        idx=index,
                        v=float(value),
                    )
                )

            if (
                series.val is not None
                and
                series.val.numRef
                is not None
            ):

                series.val.numRef.numCache = NumData(
                    formatCode="0.00%",
                    ptCount=len(values),
                    pt=numeric_points,
                )

            if series.cat is not None:

                formula = None

                if (
                    series.cat.numRef
                    is not None
                ):
                    formula = (
                        series.cat
                        .numRef
                        .f
                    )

                elif (
                    series.cat.strRef
                    is not None
                ):
                    formula = (
                        series.cat
                        .strRef
                        .f
                    )

                if formula:

                    series.cat.numRef = None

                    series.cat.strRef = StrRef(
                        f=formula,
                        strCache=StrData(
                            ptCount=len(
                                categories
                            ),
                            pt=[
                                StrVal(
                                    idx=index,
                                    v=str(value),
                                )
                                for index, value
                                in enumerate(
                                    categories
                                )
                            ],
                        ),
                    )

            if (
                series.tx is not None
                and
                series.tx.strRef
                is not None
            ):

                series.tx.strRef.strCache = StrData(
                    ptCount=1,
                    pt=[
                        StrVal(
                            idx=0,
                            v=str(title),
                        )
                    ],
                )

    except Exception:
        # Cache hatası grafiğin oluşturulmasını engellemesin.
        pass


# ============================================================
# ÇİZGİ + NOKTA AYNI RENK
# ============================================================

def _set_series_color(
    series,
    color,
    marker="circle",
    marker_size=6,
):

    try:

        # Çizgi
        series.graphicalProperties.line.solidFill = (
            color
        )

        series.graphicalProperties.line.width = (
            22000
        )

        # Marker
        series.marker.symbol = (
            marker
        )

        series.marker.size = (
            marker_size
        )

        # Marker içi
        series.marker.graphicalProperties.solidFill = (
            color
        )

        # Marker kenarı
        series.marker.graphicalProperties.line.solidFill = (
            color
        )

    except Exception:
        pass


# ============================================================
# GÜNCEL KURLAR
# ============================================================

def _build_current_sheet(
    wb,
    latest_run_at,
    latest_rows,
):

    ws = wb.create_sheet(
        "GUNCEL_KURLAR"
    )

    ws.freeze_panes = "D2"
    ws.sheet_view.showGridLines = False

    headers = [
        "Tarih",
        "Saat",
        "Kurum / Sağlayıcı",

        "Dolar Alış",
        "Dolar Satış",
        "Dolar Makas",
        "Dolar Makas %",

        "Euro Alış",
        "Euro Satış",
        "Euro Makas",
        "Euro Makas %",

        "Gram Altın Alış",
        "Gram Altın Satış",
        "Gram Altın Makas",
        "Gram Altın Makas %",
    ]

    ws.append(
        headers
    )

    _style_header(
        ws,
        1,
        1,
        len(headers),
    )

    # DOLAR
    for col in range(
        4,
        8,
    ):
        ws.cell(
            1,
            col,
        ).fill = USD_HEADER_FILL

    # EURO
    for col in range(
        8,
        12,
    ):
        ws.cell(
            1,
            col,
        ).fill = EUR_HEADER_FILL

    # ALTIN
    for col in range(
        12,
        16,
    ):
        ws.cell(
            1,
            col,
        ).fill = XAU_HEADER_FILL

    provider_map = _provider_map(
        latest_rows
    )

    run_dt = _parse_dt(
        latest_run_at
    )

    # --------------------------------------------------------
    # ÖNEMLİ:
    #
    # Burada TARGET_BANKS filtresi YOK.
    #
    # Yani:
    # Hepsipay
    # Papara
    # Kapalıçarşı
    # Harem
    # vs.
    #
    # son run'da varsa GUNCEL_KURLAR'da görünür.
    # --------------------------------------------------------

    providers = sorted(
        provider_map.keys(),
        key=str.casefold,
    )

    for excel_row, provider in enumerate(
        providers,
        start=2,
    ):

        row_map = (
            provider_map[
                provider
            ]
        )

        # Tarih
        if run_dt:
            ws.cell(
                excel_row,
                1,
                run_dt.date(),
            )

        # Saat
        if run_dt:
            ws.cell(
                excel_row,
                2,
                run_dt.time(),
            )

        # Sağlayıcı
        provider_cell = ws.cell(
            excel_row,
            3,
            provider,
        )

        provider_cell.fill = (
            _provider_fill(
                provider
            )
        )

        provider_cell.font = Font(
            bold=True
        )

        layout = {
            "USD": (
                4,
                5,
                6,
                7,
            ),
            "EUR": (
                8,
                9,
                10,
                11,
            ),
            "XAU": (
                12,
                13,
                14,
                15,
            ),
        }

        for (
            code,
            columns,
        ) in layout.items():

            (
                buy_col,
                sell_col,
                spread_col,
                pct_col,
            ) = columns

            fill = _product_fill(
                code
            )

            # Ürün grubu renklendir.
            for col in columns:
                ws.cell(
                    excel_row,
                    col,
                ).fill = fill

            item = row_map.get(
                code
            )

            # Sağlayıcı bu ürünü sunmuyorsa boş bırak.
            if not item:
                continue

            (
                buy,
                sell,
                spread,
                spread_pct,
            ) = _get_spread_values(
                item
            )

            if buy is not None:
                ws.cell(
                    excel_row,
                    buy_col,
                    buy,
                )

            if sell is not None:
                ws.cell(
                    excel_row,
                    sell_col,
                    sell,
                )

            if spread is not None:
                ws.cell(
                    excel_row,
                    spread_col,
                    spread,
                )

            if spread_pct is not None:
                ws.cell(
                    excel_row,
                    pct_col,
                    spread_pct / 100,
                )

        # Formatlar
        ws.cell(
            excel_row,
            1,
        ).number_format = "dd.mm.yyyy"

        ws.cell(
            excel_row,
            2,
        ).number_format = "hh:mm:ss"

        for col in (
            4,
            5,
            6,
            8,
            9,
            10,
            12,
            13,
            14,
        ):
            ws.cell(
                excel_row,
                col,
            ).number_format = (
                "#,##0.0000"
            )

        for col in (
            7,
            11,
            15,
        ):
            ws.cell(
                excel_row,
                col,
            ).number_format = (
                "0.00%"
            )

    # Tablo
    if ws.max_row >= 2:

        _apply_table(
            ws,
            1,
            ws.max_row,
            len(headers),
            "GuncelKurlarTable",
        )

        # En düşük makas conditional formatting
        for col_letter in (
            "G",
            "K",
            "O",
        ):

            formula = (
                f'AND('
                f'{col_letter}2<>"",'
                f'{col_letter}2='
                f'MIN('
                f'${col_letter}$2:'
                f'${col_letter}${ws.max_row}'
                f')'
                f')'
            )

            rule = FormulaRule(
                formula=[
                    formula
                ],
                fill=BEST_FILL,
                font=BEST_FONT,
            )

            ws.conditional_formatting.add(
                (
                    f"{col_letter}2:"
                    f"{col_letter}"
                    f"{ws.max_row}"
                ),
                rule,
            )

    _set_widths(
        ws,
        {
            "A": 13,
            "B": 12,
            "C": 26,

            "D": 15,
            "E": 15,
            "F": 15,
            "G": 14,

            "H": 15,
            "I": 15,
            "J": 15,
            "K": 14,

            "L": 18,
            "M": 18,
            "N": 18,
            "O": 17,
        },
    )


# ============================================================
# GEÇMİŞ
# ============================================================

def _build_history_sheet(
    wb,
    history,
):

    ws = wb.create_sheet(
        "GECMIS"
    )

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    headers = [
        "Tarih",
        "Saat",
        "Kurum / Sağlayıcı",
        "Ürün",
        "Alış",
        "Satış",
        "Makas",
        "Makas %",
        "Durum",
        "Kaynak",
        "Sitedeki Makas",
        "Sitedeki Makas %",
        "Not",
        "Gerçek Ürün Çekim Saati",
    ]

    ws.append(
        headers
    )

    _style_header(
        ws,
        1,
        1,
        len(headers),
    )

    sorted_history = sorted(
        history,
        key=_history_sort_key,
    )

    for excel_row, item in enumerate(
        sorted_history,
        start=2,
    ):

        run_dt = _parse_dt(
            item.get("run_at")
        )

        scraped_dt = _parse_dt(
            item.get("scraped_at")
        )

        code = (
            item.get("code")
            or ""
        ).strip().upper()

        (
            buy,
            sell,
            spread,
            pct,
        ) = _get_spread_values(
            item
        )

        site_spread = _to_float(
            item.get(
                "site_spread"
            )
        )

        site_pct = _to_float(
            item.get(
                "site_spread_pct"
            )
        )

        raw_status = (
            item.get(
                "status",
                "",
            )
        )

        # Tarih
        if run_dt:
            ws.cell(
                excel_row,
                1,
                run_dt.date(),
            )

        # Saat
        if run_dt:
            ws.cell(
                excel_row,
                2,
                run_dt.time(),
            )

        # Provider
        provider = (
            item.get(
                "provider",
                "",
            )
        )

        provider_cell = ws.cell(
            excel_row,
            3,
            provider,
        )

        provider_cell.fill = (
            _provider_fill(
                provider
            )
        )

        provider_cell.font = Font(
            bold=True
        )

        # Ürün
        ws.cell(
            excel_row,
            4,
            item.get(
                "product",
                PRODUCT_NAMES.get(
                    code,
                    code,
                ),
            ),
        )

        product_fill = (
            _product_fill(
                code
            )
        )

        # Ürün + değer hücreleri aynı ürün tonunda
        for col in range(
            4,
            9,
        ):
            ws.cell(
                excel_row,
                col,
            ).fill = (
                product_fill
            )

        ws.cell(
            excel_row,
            4,
        ).font = Font(
            bold=True
        )

        if buy is not None:
            ws.cell(
                excel_row,
                5,
                buy,
            )

        if sell is not None:
            ws.cell(
                excel_row,
                6,
                sell,
            )

        if spread is not None:
            ws.cell(
                excel_row,
                7,
                spread,
            )

        if pct is not None:
            ws.cell(
                excel_row,
                8,
                pct / 100,
            )

        # Durum
        ws.cell(
            excel_row,
            9,
            _display_status(
                raw_status
            ),
        )

        # Kaynak
        ws.cell(
            excel_row,
            10,
            item.get(
                "source_url",
                "",
            ),
        )

        # Sitedeki makas
        if site_spread is not None:
            ws.cell(
                excel_row,
                11,
                site_spread,
            )

        if site_pct is not None:
            ws.cell(
                excel_row,
                12,
                site_pct / 100,
            )

        # Not
        ws.cell(
            excel_row,
            13,
            item.get(
                "note",
                "",
            ),
        )

        # Gerçek ürün çekim saati
        if scraped_dt:
            ws.cell(
                excel_row,
                14,
                scraped_dt.time(),
            )

        # ----------------------------------------------------
        # HAFTA SONU
        #
        # GECMIS sayfasında hafta sonu tarih hücresi sarı.
        # ----------------------------------------------------

        if (
            run_dt
            and
            run_dt.weekday()
            in (
                5,
                6,
            )
        ):

            date_cell = ws.cell(
                excel_row,
                1,
            )

            date_cell.fill = (
                WEEKEND_FILL
            )

            date_cell.font = Font(
                bold=True,
                color="9C6500",
            )

        # Formatlar
        ws.cell(
            excel_row,
            1,
        ).number_format = (
            "dd.mm.yyyy"
        )

        ws.cell(
            excel_row,
            2,
        ).number_format = (
            "hh:mm:ss"
        )

        ws.cell(
            excel_row,
            14,
        ).number_format = (
            "hh:mm:ss"
        )

        for col in (
            5,
            6,
            7,
            11,
        ):
            ws.cell(
                excel_row,
                col,
            ).number_format = (
                "#,##0.0000"
            )

        for col in (
            8,
            12,
        ):
            ws.cell(
                excel_row,
                col,
            ).number_format = (
                "0.00%"
            )

        # Kaynak linki
        source_cell = ws.cell(
            excel_row,
            10,
        )

        if source_cell.value:

            source_cell.hyperlink = (
                source_cell.value
            )

            source_cell.style = (
                "Hyperlink"
            )

        # Durum rengi
        status_cell = ws.cell(
            excel_row,
            9,
        )

        status_cell.fill = (
            _status_fill(
                str(
                    raw_status
                )
            )
        )

        status_cell.font = Font(
            bold=True
        )

    if ws.max_row >= 2:

        _apply_table(
            ws,
            1,
            ws.max_row,
            len(headers),
            "GecmisTable",
        )

    _set_widths(
        ws,
        {
            "A": 13,
            "B": 12,
            "C": 26,
            "D": 16,
            "E": 15,
            "F": 15,
            "G": 15,
            "H": 14,
            "I": 20,
            "J": 45,
            "K": 17,
            "L": 18,
            "M": 60,
            "N": 22,
        },
    )


# ============================================================
# TREND VERİLERİ
# ============================================================

def _build_bank_trends(
    history,
):

    bank_trends = {}

    for row in history:

        provider = (
            row.get("provider")
            or ""
        ).strip()

        code = (
            row.get("code")
            or ""
        ).strip().upper()

        run_at = (
            row.get("run_at")
        )

        run_dt = _parse_dt(
            run_at
        )

        # Özet grafikler sadece 5 banka
        if provider not in TARGET_BANKS:
            continue

        if code not in PRODUCT_ORDER:
            continue

        if (
            not run_at
            or not run_dt
        ):
            continue

        if (
            row.get("status")
            == "ERROR"
        ):
            continue

        (
            _buy,
            _sell,
            _spread,
            pct,
        ) = _get_spread_values(
            row
        )

        if pct is None:
            continue

        bucket = (
            bank_trends.setdefault(
                run_at,
                {
                    "dt": run_dt,
                    "banks": {},
                },
            )
        )

        bucket[
            "banks"
        ].setdefault(
            provider,
            {},
        )[code] = (
            pct / 100
        )

    return sorted(
        bank_trends.values(),
        key=lambda x: x["dt"],
    )


# ============================================================
# AYLIK ORTALAMA
# ============================================================

def _monthly_averages(
    history,
):

    buckets = {}

    for row in history:

        provider = (
            row.get("provider")
            or ""
        ).strip()

        code = (
            row.get("code")
            or ""
        ).strip().upper()

        dt = _parse_dt(
            row.get("run_at")
        )

        if provider not in TARGET_BANKS:
            continue

        if code not in PRODUCT_ORDER:
            continue

        if not dt:
            continue

        if (
            row.get("status")
            == "ERROR"
        ):
            continue

        (
            _buy,
            _sell,
            _spread,
            pct,
        ) = _get_spread_values(
            row
        )

        if pct is None:
            continue

        key = (
            dt.year,
            dt.month,
            provider,
            code,
        )

        buckets.setdefault(
            key,
            [],
        ).append(
            pct / 100
        )

    result = {}

    for (
        year,
        month,
        provider,
        code,
    ), values in buckets.items():

        result.setdefault(
            (
                year,
                month,
            ),
            {},
        ).setdefault(
            provider,
            {},
        )[code] = (
            sum(values)
            / len(values)
        )

    return result


# ============================================================
# GRAFİK ORTAK STİL
# ============================================================

def _style_chart(
    chart,
    title,
    legend=False,
):

    chart.style = 10
    chart.title = title

    chart.height = 8.5
    chart.width = 21.0

    # --------------------------------------------------------
    # KRİTİK DÜZELTME
    #
    # Excel gizli veri davranışı nedeniyle grafiklerin
    # kaybolmaması için bunu False yapıyoruz.
    # --------------------------------------------------------

    try:
        chart.visible_cells_only = False
    except Exception:
        pass

    chart.y_axis.title = (
        "Makas %"
    )

    chart.x_axis.title = None

    try:

        chart.x_axis.axPos = "b"
        chart.x_axis.delete = False
        chart.x_axis.tickLblPos = "low"

        chart.x_axis.tickLblSkip = 1
        chart.x_axis.tickMarkSkip = 1

        chart.x_axis.majorTickMark = (
            "none"
        )

        chart.x_axis.minorTickMark = (
            "none"
        )

        chart.y_axis.majorGridlines = (
            None
        )

        chart.y_axis.majorTickMark = (
            "none"
        )

        chart.y_axis.minorTickMark = (
            "none"
        )

        chart.y_axis.numFmt = (
            "0.00%"
        )

    except Exception:
        pass

    if legend:

        try:

            chart.legend.position = (
                "b"
            )

            chart.legend.overlay = (
                False
            )

        except Exception:
            pass

    else:

        chart.legend = None

    # Tarih yazıları küçük ve eğik.
    try:

        chart.x_axis.txPr = RichText(
            bodyPr=RichTextProperties(
                rot=-2700000
            ),
            p=[
                Paragraph(
                    pPr=ParagraphProperties(
                        defRPr=CharacterProperties(
                            sz=650
                        )
                    ),
                    endParaRPr=CharacterProperties(
                        sz=650
                    ),
                )
            ],
        )

    except Exception:
        pass


# ============================================================
# ÖZET
# ============================================================

def _build_summary_sheet(
    wb,
    latest_run_at,
    latest_rows,
    history,
):

    ws = wb.create_sheet(
        "OZET"
    )

    ws.sheet_view.showGridLines = (
        False
    )

    # ========================================================
    # BAŞLIK
    # ========================================================

    ws.merge_cells(
        "A1:F1"
    )

    ws["A1"] = (
        "Döviz ve Altın Kur Takip Özeti"
    )

    ws["A1"].fill = (
        TITLE_FILL
    )

    ws["A1"].font = Font(
        bold=True,
        size=16,
        color="1F4E78",
    )

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    # ========================================================
    # GÜNLÜK TREND
    # ========================================================

    trend_runs = (
        _build_bank_trends(
            history
        )
    )

    category_values = [
        _date_label(
            run["dt"],
            include_time=False,
        )
        for run in trend_runs
    ]

    # ========================================================
    # TOPLU GÜNLÜK GRAFİK HELPER
    # ========================================================

    daily_helper_row = 2

    daily_date_col = (
        HELPER_START_COL
    )

    ws.cell(
        daily_helper_row,
        daily_date_col,
        "Tarih",
    )

    daily_columns = {}

    current_col = (
        daily_date_col
        + 1
    )

    for code in (
        "USD",
        "EUR",
        "XAU",
    ):

        daily_columns[
            code
        ] = {}

        for bank in TARGET_BANKS:

            daily_columns[
                code
            ][bank] = (
                current_col
            )

            ws.cell(
                daily_helper_row,
                current_col,
                bank,
            )

            current_col += 1

    # Veri
    for (
        row_offset,
        run,
    ) in enumerate(
        trend_runs,
        start=1,
    ):

        row = (
            daily_helper_row
            + row_offset
        )

        ws.cell(
            row,
            daily_date_col,
            _date_label(
                run["dt"],
                include_time=False,
            ),
        )

        for code in (
            "USD",
            "EUR",
            "XAU",
        ):

            for bank in TARGET_BANKS:

                value = (
                    run["banks"]
                    .get(
                        bank,
                        {},
                    )
                    .get(
                        code
                    )
                )

                if value is None:
                    continue

                cell = ws.cell(
                    row,
                    daily_columns[
                        code
                    ][bank],
                    value,
                )

                cell.number_format = (
                    "0.00%"
                )

    daily_helper_end = (
        daily_helper_row
        + len(trend_runs)
    )

    # ========================================================
    # 1. ÜSTTE 5 BANKA TOPLU GRAFİKLER
    # ========================================================

    top_positions = [
        "A3",
        "J3",
        "T3",
    ]

    if trend_runs:

        for (
            product_index,
            code,
        ) in enumerate(
            (
                "USD",
                "EUR",
                "XAU",
            )
        ):

            chart = LineChart()

            _style_chart(
                chart,
                (
                    f"{PRODUCT_NAMES[code]} "
                    "- 5 Banka Makas %"
                ),
                legend=True,
            )

            cached = []

            for bank in TARGET_BANKS:

                data_col = (
                    daily_columns[
                        code
                    ][bank]
                )

                data = Reference(
                    ws,
                    min_col=data_col,
                    max_col=data_col,
                    min_row=daily_helper_row,
                    max_row=daily_helper_end,
                )

                chart.add_data(
                    data,
                    titles_from_data=True,
                )

                values = [
                    (
                        run["banks"]
                        .get(
                            bank,
                            {},
                        )
                        .get(
                            code
                        )
                    )
                    for run
                    in trend_runs
                ]

                cached.append(
                    (
                        bank,
                        values,
                    )
                )

            categories = Reference(
                ws,
                min_col=daily_date_col,
                min_row=(
                    daily_helper_row
                    + 1
                ),
                max_row=daily_helper_end,
            )

            chart.set_categories(
                categories
            )

            # -----------------------------------------------
            # HER BANKA:
            # çizgi + marker = aynı renk
            # -----------------------------------------------

            for (
                index,
                series,
            ) in enumerate(
                chart.series
            ):

                bank = (
                    TARGET_BANKS[
                        index
                    ]
                )

                _set_series_color(
                    series,
                    BANK_LINE_COLORS[
                        bank
                    ],
                    marker="circle",
                    marker_size=6,
                )

            _cache_line_chart(
                chart,
                category_values,
                cached,
            )

            ws.add_chart(
                chart,
                top_positions[
                    product_index
                ],
            )

    # ========================================================
    # 2. AYLIK ORTALAMA TABLO
    # ========================================================

    monthly = (
        _monthly_averages(
            history
        )
    )

    months = sorted(
        monthly.keys()
    )

    monthly_title_row = 26

    ws.merge_cells(
        start_row=monthly_title_row,
        start_column=1,
        end_row=monthly_title_row,
        end_column=5,
    )

    monthly_title = ws.cell(
        monthly_title_row,
        1,
        "AYLIK ORTALAMA MAKAS %",
    )

    monthly_title.fill = (
        TITLE_FILL
    )

    monthly_title.font = Font(
        bold=True,
        size=13,
        color="1F4E78",
    )

    monthly_title.alignment = (
        Alignment(
            horizontal="center",
            vertical="center",
        )
    )

    monthly_headers = [
        "Ay",
        "Banka",
        "Dolar Ort. Makas %",
        "Euro Ort. Makas %",
        "Gram Altın Ort. Makas %",
    ]

    for (
        col,
        header,
    ) in enumerate(
        monthly_headers,
        start=1,
    ):

        ws.cell(
            27,
            col,
            header,
        )

    _style_header(
        ws,
        27,
        1,
        5,
    )

    ws["C27"].fill = (
        USD_HEADER_FILL
    )

    ws["D27"].fill = (
        EUR_HEADER_FILL
    )

    ws["E27"].fill = (
        XAU_HEADER_FILL
    )

    monthly_table_row = 28

    for (
        year,
        month,
    ) in months:

        month_label = (
            f"{MONTH_NAMES[month]} "
            f"{year}"
        )

        for bank in TARGET_BANKS:

            # Ay
            ws.cell(
                monthly_table_row,
                1,
                month_label,
            )

            # Banka
            bank_cell = ws.cell(
                monthly_table_row,
                2,
                bank,
            )

            bank_cell.fill = (
                _provider_fill(
                    bank
                )
            )

            bank_cell.font = Font(
                bold=True
            )

            # Ürün ortalamaları
            for (
                col,
                code,
            ) in (
                (
                    3,
                    "USD",
                ),
                (
                    4,
                    "EUR",
                ),
                (
                    5,
                    "XAU",
                ),
            ):

                cell = ws.cell(
                    monthly_table_row,
                    col,
                )

                cell.fill = (
                    _product_fill(
                        code
                    )
                )

                value = (
                    monthly
                    .get(
                        (
                            year,
                            month,
                        ),
                        {},
                    )
                    .get(
                        bank,
                        {},
                    )
                    .get(
                        code
                    )
                )

                if value is not None:

                    cell.value = (
                        value
                    )

                    cell.number_format = (
                        "0.00%"
                    )

            for col in range(
                1,
                6,
            ):

                ws.cell(
                    monthly_table_row,
                    col,
                ).border = (
                    THIN_BORDER
                )

            monthly_table_row += 1

    # ========================================================
    # AYLIK HELPER
    # ========================================================

    monthly_helper_row = 100

    monthly_date_col = (
        HELPER_START_COL
    )

    ws.cell(
        monthly_helper_row,
        monthly_date_col,
        "Ay",
    )

    monthly_columns = {}

    current_monthly_col = (
        monthly_date_col
        + 1
    )

    for code in (
        "USD",
        "EUR",
        "XAU",
    ):

        monthly_columns[
            code
        ] = {}

        for bank in TARGET_BANKS:

            monthly_columns[
                code
            ][bank] = (
                current_monthly_col
            )

            ws.cell(
                monthly_helper_row,
                current_monthly_col,
                bank,
            )

            current_monthly_col += 1

    monthly_labels = []

    for (
        row_offset,
        month_key,
    ) in enumerate(
        months,
        start=1,
    ):

        year, month = (
            month_key
        )

        label = (
            f"{MONTH_NAMES[month]} "
            f"{year}"
        )

        monthly_labels.append(
            label
        )

        row = (
            monthly_helper_row
            + row_offset
        )

        ws.cell(
            row,
            monthly_date_col,
            label,
        )

        for code in (
            "USD",
            "EUR",
            "XAU",
        ):

            for bank in TARGET_BANKS:

                value = (
                    monthly
                    .get(
                        month_key,
                        {},
                    )
                    .get(
                        bank,
                        {},
                    )
                    .get(
                        code
                    )
                )

                if value is None:
                    continue

                cell = ws.cell(
                    row,
                    monthly_columns[
                        code
                    ][bank],
                    value,
                )

                cell.number_format = (
                    "0.00%"
                )

    monthly_helper_end = (
        monthly_helper_row
        + len(months)
    )

    # ========================================================
    # 3. AYLIK ORTALAMA GRAFİKLERİ
    # ========================================================

    monthly_positions = [
        "A42",
        "J42",
        "T42",
    ]

    if months:

        for (
            product_index,
            code,
        ) in enumerate(
            (
                "USD",
                "EUR",
                "XAU",
            )
        ):

            chart = LineChart()

            _style_chart(
                chart,
                (
                    f"{PRODUCT_NAMES[code]} "
                    "- Aylık Ortalama Makas %"
                ),
                legend=True,
            )

            cached = []

            for bank in TARGET_BANKS:

                data_col = (
                    monthly_columns[
                        code
                    ][bank]
                )

                data = Reference(
                    ws,
                    min_col=data_col,
                    max_col=data_col,
                    min_row=monthly_helper_row,
                    max_row=monthly_helper_end,
                )

                chart.add_data(
                    data,
                    titles_from_data=True,
                )

                values = [
                    (
                        monthly
                        .get(
                            month_key,
                            {},
                        )
                        .get(
                            bank,
                            {},
                        )
                        .get(
                            code
                        )
                    )
                    for month_key
                    in months
                ]

                cached.append(
                    (
                        bank,
                        values,
                    )
                )

            categories = Reference(
                ws,
                min_col=monthly_date_col,
                min_row=(
                    monthly_helper_row
                    + 1
                ),
                max_row=monthly_helper_end,
            )

            chart.set_categories(
                categories
            )

            # -----------------------------------------------
            # AYLIK GRAFİK:
            #
            # ÇİZGİ VE NOKTA AYNI BANKA RENGİ
            # -----------------------------------------------

            for (
                index,
                series,
            ) in enumerate(
                chart.series
            ):

                bank = (
                    TARGET_BANKS[
                        index
                    ]
                )

                _set_series_color(
                    series,
                    BANK_LINE_COLORS[
                        bank
                    ],
                    marker="circle",
                    marker_size=7,
                )

            _cache_line_chart(
                chart,
                monthly_labels,
                cached,
            )

            ws.add_chart(
                chart,
                monthly_positions[
                    product_index
                ],
            )

    # ========================================================
    # 4. 15 TEKİL BANKA GRAFİĞİ
    # ========================================================

    chart_columns = [
        "A",
        "J",
        "T",
    ]

    chart_row_starts = [
        65,
        88,
        111,
        134,
        157,
    ]

    individual_helper_row = 200

    individual_start_col = (
        HELPER_START_COL
    )

    bank_helper_columns = {}

    current_individual_col = (
        individual_start_col
    )

    for bank in TARGET_BANKS:

        bank_helper_columns[
            bank
        ] = {
            "DATE": (
                current_individual_col
            ),
            "USD": (
                current_individual_col
                + 1
            ),
            "EUR": (
                current_individual_col
                + 2
            ),
            "XAU": (
                current_individual_col
                + 3
            ),
        }

        ws.cell(
            individual_helper_row,
            current_individual_col,
            f"{bank} Tarih",
        )

        ws.cell(
            individual_helper_row,
            current_individual_col + 1,
            "DOLAR",
        )

        ws.cell(
            individual_helper_row,
            current_individual_col + 2,
            "EURO",
        )

        ws.cell(
            individual_helper_row,
            current_individual_col + 3,
            "GRAM ALTIN",
        )

        current_individual_col += 4

    for (
        row_offset,
        run,
    ) in enumerate(
        trend_runs,
        start=1,
    ):

        row = (
            individual_helper_row
            + row_offset
        )

        for bank in TARGET_BANKS:

            cols = (
                bank_helper_columns[
                    bank
                ]
            )

            ws.cell(
                row,
                cols["DATE"],
                _date_label(
                    run["dt"],
                    include_time=True,
                ),
            )

            for code in (
                "USD",
                "EUR",
                "XAU",
            ):

                value = (
                    run["banks"]
                    .get(
                        bank,
                        {},
                    )
                    .get(
                        code
                    )
                )

                if value is None:
                    continue

                cell = ws.cell(
                    row,
                    cols[
                        code
                    ],
                    value,
                )

                cell.number_format = (
                    "0.00%"
                )

    individual_helper_end = (
        individual_helper_row
        + len(trend_runs)
    )

    # --------------------------------------------------------
    # Grafikler
    # --------------------------------------------------------

    if trend_runs:

        for (
            bank_index,
            bank,
        ) in enumerate(
            TARGET_BANKS
        ):

            cols = (
                bank_helper_columns[
                    bank
                ]
            )

            individual_categories = [
                _date_label(
                    run["dt"],
                    include_time=True,
                )
                for run
                in trend_runs
            ]

            for (
                product_index,
                code,
            ) in enumerate(
                (
                    "USD",
                    "EUR",
                    "XAU",
                )
            ):

                chart = LineChart()

                _style_chart(
                    chart,
                    (
                        f"{bank} - "
                        f"{PRODUCT_NAMES[code]} "
                        "Makas %"
                    ),
                    legend=False,
                )

                # -------------------------------------------
                # Bankaya göre pastel arka plan
                # -------------------------------------------

                bank_bg = (
                    BANK_CHART_COLORS.get(
                        bank,
                        DEFAULT_BANK_CHART_COLOR,
                    )
                )

                try:

                    chart.graphical_properties = (
                        GraphicalProperties(
                            noFill=False,
                            solidFill=bank_bg,
                        )
                    )

                    chart.plot_area.graphicalProperties = (
                        GraphicalProperties(
                            noFill=False,
                            solidFill=bank_bg,
                        )
                    )

                except Exception:
                    pass

                # Veri
                data = Reference(
                    ws,
                    min_col=cols[
                        code
                    ],
                    max_col=cols[
                        code
                    ],
                    min_row=individual_helper_row,
                    max_row=individual_helper_end,
                )

                categories = Reference(
                    ws,
                    min_col=cols[
                        "DATE"
                    ],
                    min_row=(
                        individual_helper_row
                        + 1
                    ),
                    max_row=individual_helper_end,
                )

                chart.add_data(
                    data,
                    titles_from_data=True,
                )

                chart.set_categories(
                    categories
                )

                values = [
                    (
                        run["banks"]
                        .get(
                            bank,
                            {},
                        )
                        .get(
                            code
                        )
                    )
                    for run
                    in trend_runs
                ]

                # -------------------------------------------
                # ÜRÜN RENGİ
                #
                # Dolar = mavi
                # Euro = turuncu
                # Altın = sarı
                #
                # Çizgi ve marker AYNI renk.
                #
                # Hafta sonu marker rengi DEĞİŞMEZ.
                # -------------------------------------------

                if chart.series:

                    if code == "USD":
                        marker = "circle"

                    elif code == "EUR":
                        marker = "square"

                    else:
                        marker = "triangle"

                    _set_series_color(
                        chart.series[0],
                        PRODUCT_LINE_COLORS[
                            code
                        ],
                        marker=marker,
                        marker_size=6,
                    )

                # -------------------------------------------
                # Değer etiketleri
                # -------------------------------------------

                chart.dLbls = (
                    DataLabelList()
                )

                chart.dLbls.showVal = (
                    True
                )

                chart.dLbls.numFmt = (
                    "0.00%"
                )

                chart.dLbls.dLblPos = (
                    "t"
                )

                chart.dLbls.showLegendKey = (
                    False
                )

                chart.dLbls.showCatName = (
                    False
                )

                chart.dLbls.showSerName = (
                    False
                )

                _cache_line_chart(
                    chart,
                    individual_categories,
                    [
                        (
                            PRODUCT_NAMES[
                                code
                            ],
                            values,
                        )
                    ],
                )

                # -------------------------------------------
                # Y eksenini verilere yaklaştır
                # -------------------------------------------

                valid_values = [
                    value
                    for value
                    in values
                    if value is not None
                ]

                if valid_values:

                    minimum = min(
                        valid_values
                    )

                    maximum = max(
                        valid_values
                    )

                    padding = max(
                        (
                            maximum
                            - minimum
                        )
                        * 0.20,
                        maximum
                        * 0.02,
                        0.00005,
                    )

                    chart.y_axis.scaling.min = (
                        max(
                            0,
                            minimum
                            - padding,
                        )
                    )

                    chart.y_axis.scaling.max = (
                        maximum
                        + padding
                    )

                position = (
                    f"{chart_columns[product_index]}"
                    f"{chart_row_starts[bank_index]}"
                )

                ws.add_chart(
                    chart,
                    position,
                )

    # ========================================================
    # 5. SON VERİ ÇEKİMİ ÖZETİ
    # ========================================================

    run_dt = _parse_dt(
        latest_run_at
    )

    providers = {
        (
            row.get("provider")
            or ""
        ).strip()
        for row in latest_rows
        if (
            row.get("provider")
            or ""
        ).strip()
    }

    summary_title_row = 183

    ws.merge_cells(
        start_row=summary_title_row,
        start_column=1,
        end_row=summary_title_row,
        end_column=4,
    )

    summary_title = ws.cell(
        summary_title_row,
        1,
        "SON VERİ ÇEKİMİ ÖZETİ",
    )

    summary_title.fill = (
        TITLE_FILL
    )

    summary_title.font = Font(
        bold=True,
        size=12,
        color="1F4E78",
    )

    summary_title.alignment = (
        Alignment(
            horizontal="center",
            vertical="center",
        )
    )

    summary_rows = [
        (
            "Veri tarihi",
            (
                run_dt.date()
                if run_dt
                else ""
            ),
        ),
        (
            "Veri çekim saati",
            (
                run_dt.time()
                if run_dt
                else ""
            ),
        ),
        (
            "Son çekimde bulunan sağlayıcı sayısı",
            len(
                providers
            ),
        ),
        (
            "Son çekimde alınan toplam ürün kaydı",
            len(
                latest_rows
            ),
        ),
        (
            "Hatalı kayıt sayısı",
            sum(
                1
                for row
                in latest_rows
                if (
                    row.get("status")
                    == "ERROR"
                )
            ),
        ),
        (
            "Kontrol gereken kayıt sayısı",
            sum(
                1
                for row
                in latest_rows
                if (
                    row.get("status")
                    == "KONTROL"
                )
            ),
        ),
    ]

    row = (
        summary_title_row
        + 1
    )

    for (
        label,
        value,
    ) in summary_rows:

        label_cell = ws.cell(
            row,
            1,
            label,
        )

        value_cell = ws.cell(
            row,
            2,
            value,
        )

        label_cell.font = Font(
            bold=True
        )

        label_cell.fill = PatternFill(
            "solid",
            fgColor="EAF2F8",
        )

        label_cell.border = (
            THIN_BORDER
        )

        value_cell.border = (
            THIN_BORDER
        )

        row += 1

    ws.cell(
        summary_title_row + 1,
        2,
    ).number_format = (
        "dd.mm.yyyy"
    )

    ws.cell(
        summary_title_row + 2,
        2,
    ).number_format = (
        "hh:mm:ss"
    )

    # ========================================================
    # 6. EN DÜŞÜK MAKAS
    #
    # SADECE 5 HEDEF BANKA
    # ========================================================

    best_title_row = 191

    ws.merge_cells(
        start_row=best_title_row,
        start_column=1,
        end_row=best_title_row,
        end_column=6,
    )

    best_title = ws.cell(
        best_title_row,
        1,
        "5 BANKA ARASINDA EN DÜŞÜK GÜNCEL MAKAS",
    )

    best_title.fill = (
        TITLE_FILL
    )

    best_title.font = Font(
        bold=True,
        size=12,
        color="1F4E78",
    )

    best_title.alignment = (
        Alignment(
            horizontal="center",
            vertical="center",
        )
    )

    best_header_row = (
        best_title_row
        + 1
    )

    best_headers = [
        "Ürün",
        "Verisi Bulunan Banka Sayısı",
        "En Düşük Makas %",
        "En Avantajlı Banka",
        "Alış",
        "Satış",
    ]

    for (
        col,
        header,
    ) in enumerate(
        best_headers,
        start=1,
    ):

        ws.cell(
            best_header_row,
            col,
            header,
        )

    _style_header(
        ws,
        best_header_row,
        1,
        6,
    )

    best_row = (
        best_header_row
        + 1
    )

    for code in (
        "USD",
        "EUR",
        "XAU",
    ):

        valid = []

        for item in latest_rows:

            provider = (
                item.get("provider")
                or ""
            ).strip()

            item_code = (
                item.get("code")
                or ""
            ).strip().upper()

            if (
                provider
                not in TARGET_BANKS
            ):
                continue

            if item_code != code:
                continue

            if (
                item.get("status")
                == "ERROR"
            ):
                continue

            (
                buy,
                sell,
                _spread,
                pct,
            ) = _get_spread_values(
                item
            )

            if (
                pct is None
                or buy is None
                or sell is None
            ):
                continue

            valid.append(
                (
                    pct,
                    provider,
                    buy,
                    sell,
                )
            )

        best = None

        if valid:

            best = min(
                valid,
                key=lambda x: x[0],
            )

        product_cell = ws.cell(
            best_row,
            1,
            PRODUCT_NAMES[
                code
            ],
        )

        product_cell.fill = (
            _product_fill(
                code
            )
        )

        product_cell.font = Font(
            bold=True
        )

        # Kaç bankanın geçerli verisi var?
        ws.cell(
            best_row,
            2,
            len(valid),
        )

        if best:

            (
                pct,
                provider,
                buy,
                sell,
            ) = best

            ws.cell(
                best_row,
                3,
                pct / 100,
            ).number_format = (
                "0.00%"
            )

            bank_cell = ws.cell(
                best_row,
                4,
                provider,
            )

            bank_cell.fill = (
                _provider_fill(
                    provider
                )
            )

            bank_cell.font = Font(
                bold=True
            )

            ws.cell(
                best_row,
                5,
                buy,
            ).number_format = (
                "#,##0.0000"
            )

            ws.cell(
                best_row,
                6,
                sell,
            ).number_format = (
                "#,##0.0000"
            )

        for col in range(
            1,
            7,
        ):

            ws.cell(
                best_row,
                col,
            ).border = (
                THIN_BORDER
            )

        best_row += 1

    # ========================================================
    # ÖZET SAYFASI GENİŞLİKLERİ
    # ========================================================

    _set_widths(
        ws,
        {
            "A": 30,
            "B": 23,
            "C": 20,
            "D": 23,
            "E": 18,
            "F": 18,

            "G": 16,
            "H": 16,
            "I": 16,

            "J": 16,
            "K": 16,
            "L": 16,
            "M": 16,
            "N": 16,
            "O": 16,
            "P": 16,
            "Q": 16,
            "R": 16,
            "S": 16,
            "T": 16,
        },
    )

    # --------------------------------------------------------
    # DİKKAT:
    #
    # AA ve sonrasındaki helper sütunlarını GİZLEMİYORUZ.
    #
    # Bu bilinçli.
    #
    # Grafiklerin kaybolmasına neden olan problem buydu.
    # --------------------------------------------------------


# ============================================================
# ANA EXCEL FONKSİYONU
# ============================================================

def build_excel(
    history_path: str | Path,
    output_path: str | Path,
) -> None:

    raw_history = (
        read_history(
            history_path
        )
    )

    if not raw_history:

        raise RuntimeError(
            "Excel oluşturmak için geçmiş veri bulunamadı."
        )

    # ========================================================
    # GUNCEL_KURLAR
    #
    # Gerçek EN SON run.
    #
    # Böylece Papara / Hepsipay gibi sağlayıcılar
    # son run'da varsa güncel sayfada görünür.
    # ========================================================

    (
        latest_run_at,
        latest_rows,
    ) = _latest_run_rows(
        raw_history
    )

    if (
        not latest_run_at
        or not latest_rows
    ):

        raise RuntimeError(
            "Son çalıştırmaya ait geçerli veri bulunamadı."
        )

    # ========================================================
    # GECMIS + OZET
    #
    # Her günün İLK run'ı.
    #
    # Aynı gün 3 kere çalışsa bile grafiklerde/günlük geçmişte
    # yalnızca ilk çekim kullanılır.
    # ========================================================

    daily_history = (
        _daily_history(
            raw_history
        )
    )

    if not daily_history:

        raise RuntimeError(
            "Geçerli günlük geçmiş veri bulunamadı."
        )

    # ========================================================
    # WORKBOOK
    # ========================================================

    wb = Workbook()

    default_sheet = (
        wb.active
    )

    wb.remove(
        default_sheet
    )

    # --------------------------------------------------------
    # 1. GUNCEL_KURLAR
    # --------------------------------------------------------

    _build_current_sheet(
        wb,
        latest_run_at,
        latest_rows,
    )

    # --------------------------------------------------------
    # 2. GECMIS
    # --------------------------------------------------------

    _build_history_sheet(
        wb,
        daily_history,
    )

    # --------------------------------------------------------
    # 3. OZET
    # --------------------------------------------------------

    _build_summary_sheet(
        wb,
        latest_run_at,
        latest_rows,
        daily_history,
    )

    # İlk açılan sayfa GUNCEL_KURLAR
    wb.active = 0

    output_path = Path(
        output_path
    )

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    wb.save(
        output_path
    )


# ============================================================
# UYUMLULUK ALIASLARI
#
# main.py eski isimlerden birini kullanıyorsa bozulmasın.
# ============================================================

def write_excel(
    history_path: str | Path,
    output_path: str | Path,
) -> None:

    return build_excel(
        history_path,
        output_path,
    )


def create_excel(
    history_path: str | Path,
    output_path: str | Path,
) -> None:

    return build_excel(
        history_path,
        output_path,
    )


def create_excel_report(
    history_path: str | Path,
    output_path: str | Path,
) -> None:

    return build_excel(
        history_path,
        output_path,
    )
